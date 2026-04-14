"""
Decathlon Product Lookup
Improvements over v2:
  - Variation mapping: auto-detects size / color / color+size per model group
  - Short description: 2-3 bullet points generated from product fields (no API needed)
    OR via Groq when AI mode is on (better quality)
  - Performance: keyword_match_category vectorised with numpy instead of df.apply per row
  - Short description written to template's short_description column
"""

import os, io, re, json, asyncio
import numpy as np
import streamlit as st
import pandas as pd
import requests
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

try:
    from groq import AsyncGroq, Groq as SyncGroq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Decathlon Product Lookup", page_icon="🏅", layout="wide")
st.markdown("""
<style>
h1 { color: #0082C3; }
.tag {
    display:inline-block; background:#0082C3; color:white;
    border-radius:4px; padding:2px 8px; font-size:12px; margin:2px;
}
.ai-badge {
    display:inline-block; background:linear-gradient(90deg,#f55036,#ff8c00);
    color:white; border-radius:12px; padding:2px 10px;
    font-size:11px; font-weight:700; margin-left:6px;
}
.kw-badge {
    display:inline-block; background:#0082C3; color:white;
    border-radius:12px; padding:2px 10px;
    font-size:11px; font-weight:700; margin-left:6px;
}
</style>
""", unsafe_allow_html=True)

st.title("🏅 Decathlon Product Lookup")
st.markdown("Search by model number or product name — view details, images, and **download a filled upload template**.")

# ── Constants ─────────────────────────────────────────────────────────────────
IMAGE_COLS    = ["OG_image"] + [f"picture_{i}" for i in range(1, 11)]
TEMPLATE_PATH = "product-creation-template.xlsx"
DECA_CAT_PATH = "deca_cat.xlsx"
MASTER_PATH   = "Decathlon_Working_File_Split.csv"

MASTER_TO_TEMPLATE = {
    "product_name":   "Name",
    "designed_for":   "Description",
    "sku_num_sku_r3": "SellerSKU",
    "model_code":     "ParentSKU",
    "brand_name":     "Brand",
    "bar_code":       "GTIN_Barcode",
    "color":          "color",
    "model_label":    "model",
    "keywords":       "note",
    "weight":         "product_weight",
    "OG_image":       "MainImage",
    "picture_1":      "Image2",
    "picture_2":      "Image3",
    "picture_3":      "Image4",
    "picture_4":      "Image5",
    "picture_5":      "Image6",
    "picture_6":      "Image7",
    "picture_7":      "Image8",
}

CATEGORY_MATCH_FIELDS = [
    "family", "type", "department_label", "nature_label",
    "proposed_brand_name", "brand_name", "color", "channable_gender",
    "size", "keywords", "designed_for", "business_weight", "product_name",
]

GROQ_SYSTEM_CAT = """You are a product categorization expert for a sports retailer.
Given a product description and candidate category paths, pick the {top_n} best matches.
Consider brand, product type, gender, sport, and age group.

Respond with JSON only:
{{
  "categories": [
    {{"category": "<full path>", "score": 0.95}},
    ...
  ]
}}

Rules:
- Return exactly {top_n} categories ordered by confidence descending
- Only pick from the provided candidate list - never invent categories
- Scores are floats 0.0-1.0
- JSON only, nothing else"""

GROQ_SYSTEM_DESC = """You are a product copywriter for a sports retail marketplace.
Given product details, write exactly 3 short bullet points (each max 15 words) that highlight
the key features a buyer cares about. Focus on: sport/use-case, key benefit or material, target user.
Do NOT start with "Our team" or "Our designers". Be specific — mention the product name or sport.
Respond with JSON only:
{{"bullets": ["bullet 1", "bullet 2", "bullet 3"]}}
JSON only, nothing else."""

# =============================================================================
# DATA LOADING
# =============================================================================

@st.cache_data(show_spinner=False)
def load_reference_data(file_bytes: bytes):
    wb_bytes = io.BytesIO(file_bytes)
    df_cat = pd.read_excel(wb_bytes, sheet_name="category", dtype=str)
    df_cat.columns = [c.strip() for c in df_cat.columns]
    df_cat = df_cat[df_cat["export_category"].notna() & (df_cat["export_category"].str.strip() != "")]
    df_cat["export_category"]     = df_cat["export_category"].str.strip()
    df_cat["category_name_lower"] = df_cat["category_name"].str.lower().str.strip()
    df_cat["Category Path lower"] = df_cat["Category Path"].str.lower().fillna("")
    # Pre-tokenise paths for fast numpy matching
    df_cat["_path_tokens"] = df_cat["Category Path lower"].apply(
        lambda p: set(re.findall(r"[a-z]+", p))
    )
    wb_bytes.seek(0)
    df_brands = pd.read_excel(wb_bytes, sheet_name="brands", dtype=str, header=0)
    df_brands.columns = ["brand_entry"]
    df_brands = df_brands[df_brands["brand_entry"].notna()].copy()
    df_brands["brand_entry"]      = df_brands["brand_entry"].str.strip()
    df_brands["brand_name_lower"] = (
        df_brands["brand_entry"].str.split(" - ", n=1).str[-1].str.lower().str.strip()
    )
    return df_cat, df_brands


@st.cache_data(show_spinner=False)
def load_master(file_bytes: bytes, is_csv: bool) -> pd.DataFrame:
    if is_csv:
        try:
            return pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            return pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="latin-1")
    return pd.read_excel(io.BytesIO(file_bytes), dtype=str)


# =============================================================================
# TF-IDF INDEX
# =============================================================================

def _path_to_doc(path: str) -> str:
    parts = path.split(" / ")
    return " ".join(parts) + " " + " ".join(parts[-3:]) * 2


@st.cache_resource(show_spinner=False)
def build_tfidf_index(ref_bytes: bytes):
    df_cat, _ = load_reference_data(ref_bytes)
    all_paths  = df_cat["Category Path"].dropna().astype(str).tolist()
    path_set   = set(all_paths)
    leaves     = [p for p in all_paths
                  if not any(other.startswith(p + " / ") for other in path_set)]
    docs       = [_path_to_doc(p) for p in leaves]
    vectorizer = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True)
    matrix     = vectorizer.fit_transform(docs)
    path_to_export = dict(zip(df_cat["Category Path"], df_cat["export_category"]))
    return leaves, vectorizer, matrix, path_to_export


def tfidf_shortlist(queries: list, leaves, vectorizer, matrix, k: int = 30) -> list:
    qmat = vectorizer.transform(queries)
    sims = cosine_similarity(qmat, matrix)
    out  = []
    for row in sims:
        top_idx = np.argsort(row)[::-1][:k]
        out.append([leaves[i] for i in top_idx if row[i] > 0])
    return out


# =============================================================================
# KEYWORD MATCHING  — vectorised, no per-row apply()
# =============================================================================

def _build_query_string(row: pd.Series) -> str:
    parts = []
    for f in CATEGORY_MATCH_FIELDS:
        val = row.get(f, "")
        if pd.notna(val) and str(val).strip() not in ("", "-", "nan"):
            parts.append(str(val).strip().lower())
    return " ".join(parts)


@st.cache_data(show_spinner=False)
def _precompute_cat_token_arrays(df_cat_json: str):
    """Convert pre-tokenised path sets to a list for fast overlap scoring."""
    import json
    records = json.loads(df_cat_json)
    token_sets = [set(r["_path_tokens"]) for r in records]
    depths     = [r["Category Path lower"].count("/") for r in records]
    names      = [r["category_name_lower"] for r in records]
    exports    = [r["export_category"] for r in records]
    return token_sets, depths, names, exports


def keyword_match_batch(rows_df: pd.DataFrame, df_cat: pd.DataFrame) -> list:
    """
    Vectorised keyword match for ALL rows at once.
    Returns list of (primary_export, additional_export) per row.
    """
    # Build query strings for all rows
    queries = [_build_query_string(row) for _, row in rows_df.iterrows()]

    # Pre-extract category data once
    cat_token_sets = df_cat["_path_tokens"].tolist()
    cat_depths     = df_cat["Category Path lower"].str.count("/").tolist()
    cat_names      = df_cat["category_name_lower"].tolist()
    cat_exports    = df_cat["export_category"].tolist()
    n_cats         = len(cat_exports)

    results = []
    for query in queries:
        if not query:
            results.append(("", ""))
            continue
        q_tokens = set(re.findall(r"[a-z]+", query))
        # Score all categories in one list comprehension
        scores = [
            len(q_tokens & cat_token_sets[j])
            + (2 if cat_names[j] in query else 0)
            + cat_depths[j] * 0.1
            for j in range(n_cats)
        ]
        # Get top-2 indices
        top2 = sorted(range(n_cats), key=lambda j: scores[j], reverse=True)[:2]
        primary   = cat_exports[top2[0]] if scores[top2[0]] > 0 else ""
        secondary = cat_exports[top2[1]] if len(top2) > 1 and scores[top2[1]] > 0 else ""
        results.append((primary, secondary))
    return results


def keyword_match_category(row: pd.Series, df_cat: pd.DataFrame) -> tuple:
    """Single-row convenience wrapper (used for override preview)."""
    return keyword_match_batch(pd.DataFrame([row]), df_cat)[0]


# =============================================================================
# VARIATION MAPPING
# =============================================================================

def compute_variation(row: pd.Series, df_master: pd.DataFrame) -> str:
    """
    Determine the variation string for a SKU based on its model_code group.
    Logic:
      - Group all SKUs by model_code
      - If group has multiple unique colors AND multiple unique sizes → 'color,size'
      - If group has multiple unique sizes only                        → 'size'
      - If group has multiple unique colors only                       → 'color'
      - Single SKU                                                     → 'size'  (Jumia default)
    Returns the variation value string.
    """
    model_code = row.get("model_code", "")
    if not model_code or pd.isna(model_code):
        return "size"

    group = df_master[df_master["model_code"] == model_code]
    n_colors = group["color"].nunique()
    n_sizes  = group["size"].nunique()

    if n_colors > 1 and n_sizes > 1:
        return "color,size"
    elif n_colors > 1:
        return "color"
    else:
        return "size"


@st.cache_data(show_spinner=False)
def build_variation_map(master_bytes: bytes, is_csv: bool) -> dict:
    """Pre-compute variation string for every model_code → {sku: variation}."""
    df = load_master(master_bytes, is_csv)
    result = {}
    for mc, grp in df.groupby("model_code"):
        n_colors = grp["color"].nunique()
        n_sizes  = grp["size"].nunique()
        if n_colors > 1 and n_sizes > 1:
            var = "color,size"
        elif n_colors > 1:
            var = "color"
        else:
            var = "size"
        for sku in grp["sku_num_sku_r3"]:
            result[sku] = var
    return result


# =============================================================================
# SHORT DESCRIPTION  (rule-based, instant)
# =============================================================================

def _clean(val) -> str:
    if pd.isna(val) or str(val).strip() in ("", "-", "nan"):
        return ""
    return str(val).strip()


GENDER_MAP = {
    "MEN'S": "Men", "WOMEN'S": "Women", "BOYS'": "Boys", "GIRLS'": "Girls",
    "MEN": "Men", "WOMEN": "Women", "UNISEX": "Unisex", "NO GENDER": "",
    "HORSE": "",
}


def rule_based_short_desc(row: pd.Series) -> str:
    """
    Build 3 bullet points from master fields without any API call.
    Returns newline-separated bullets.
    """
    bullets = []

    # Bullet 1: Sport · Gender — use department_label (fuller than truncated type)
    dept   = _clean(row.get("department_label", "")).replace("/", "·").title()
    sport  = dept if dept else _clean(row.get("type", "")).title()
    g_raw  = _clean(row.get("channable_gender", "")).split("|")[0].strip().upper()
    gender = GENDER_MAP.get(g_raw, g_raw.title())
    if sport:
        who = f" · {gender}" if gender else ""
        bullets.append(f"• {sport}{who}")

    # Bullet 2: Key feature — skip "Our team/designers" boilerplate opener
    desc = _clean(row.get("designed_for", ""))
    if desc:
        sentences = [s.strip() for s in re.split(r"[.!?]", desc) if len(s.strip()) > 20]
        feature = next(
            (s for s in sentences if not re.match(r"our (team|design)", s, re.I)),
            sentences[0] if sentences else "",
        )
        if feature:
            trunc = feature[:120].rsplit(" ", 1)[0] if len(feature) > 120 else feature
            bullets.append(f"• {trunc}")

    # Bullet 3: Colour · Size  (strip escaped quotes and trailing junk)
    color = _clean(row.get("color", "")).split("|")[0].strip().title()
    size  = re.sub(r'"+', "", _clean(row.get("size", ""))).strip().rstrip(" .")
    if color and size and size.lower() != "no size":
        bullets.append(f"• {color} · Size {size}")
    elif color:
        bullets.append(f"• Colour: {color}")
    elif size and size.lower() != "no size":
        bullets.append(f"• Size: {size}")

    return "\n".join(bullets[:3]) if bullets else ""


# =============================================================================
# AI MATCHING  (TF-IDF → Groq, all parallel)
# =============================================================================

async def _async_rerank(idx, query, candidates, client, model, top_n, sem, task_type="cat"):
    async with sem:
        try:
            if task_type == "cat":
                cand_list = "\n".join(f"- {c}" for c in candidates)
                sys_msg   = GROQ_SYSTEM_CAT.format(top_n=top_n)
                user_msg  = f"Product: {query}\n\nCandidates:\n{cand_list}"
            else:  # desc
                sys_msg   = GROQ_SYSTEM_DESC
                user_msg  = f"Product details: {query}"

            resp = await client.chat.completions.create(
                model=model,
                temperature=0.15,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": sys_msg},
                    {"role": "user",   "content": user_msg},
                ],
            )
            raw  = resp.choices[0].message.content.strip()
            data = json.loads(raw)
            return idx, data
        except Exception as e:
            return idx, {"error": str(e)}


async def _parallel_tasks(items, client, model, sem, task_type):
    tasks = [
        _async_rerank(i, q, c, client, model, 2, sem, task_type)
        for i, (q, c) in enumerate(items)
    ]
    raw = await asyncio.gather(*tasks)
    return [r for _, r in sorted(raw, key=lambda x: x[0])]


def groq_batch(items, api_key, model, concurrency, task_type="cat"):
    async def _run():
        client = AsyncGroq(api_key=api_key)
        sem    = asyncio.Semaphore(concurrency)
        return await _parallel_tasks(items, client, model, sem, task_type)
    return asyncio.run(_run())


def ai_match_categories(rows_df, leaves, vectorizer, matrix, path_to_export,
                        api_key, model, shortlist_k=30, concurrency=10):
    queries         = [_build_query_string(row) for _, row in rows_df.iterrows()]
    candidates_list = tfidf_shortlist(queries, leaves, vectorizer, matrix, shortlist_k)
    items           = list(zip(queries, candidates_list))
    all_preds       = groq_batch(items, api_key, model, concurrency, task_type="cat")

    def _resolve(cat_path: str) -> str:
        if cat_path in path_to_export:
            return path_to_export[cat_path]
        for p, ex in path_to_export.items():
            if p.endswith(cat_path) or cat_path.endswith(p):
                return ex
        return cat_path

    results = []
    for data in all_preds:
        cats = data.get("categories", [])
        primary   = _resolve(cats[0]["category"]) if len(cats) > 0 else ""
        secondary = _resolve(cats[1]["category"]) if len(cats) > 1 else ""
        results.append((primary, secondary))
    return results


def _build_desc_query_per_model(group_df: pd.DataFrame) -> str:
    """One Groq query per model group — uses shared fields only, excludes color/size."""
    row   = group_df.iloc[0]
    parts = [
        _clean(row.get("product_name", "")),
        _clean(row.get("department_label", "")),
        _clean(row.get("brand_name", "")),
        _clean(row.get("channable_gender", "")).split("|")[0].strip(),
        _clean(row.get("designed_for", ""))[:300],
        _clean(row.get("keywords", ""))[:100],
    ]
    return " | ".join(p for p in parts if p)


def ai_short_descriptions(rows_df, api_key, model, concurrency=10):
    """
    Generate short descriptions via Groq, deduplicated per model_code.
    ~5x fewer API calls since all SKUs of a model share the same description.
    """
    # Step 1: one query per unique model_code
    model_queries: dict = {}
    model_repr:    dict = {}
    for i, (_, row) in enumerate(rows_df.iterrows()):
        mc = str(row.get("model_code", "")).strip()
        if mc and mc not in model_queries:
            group = rows_df[rows_df["model_code"] == mc]
            model_queries[mc] = _build_desc_query_per_model(group)
            model_repr[mc]    = i

    unique_models = list(model_queries.keys())
    items         = [(model_queries[mc], []) for mc in unique_models]

    # Step 2: single Groq batch for all unique models
    raw_results = groq_batch(items, api_key, model, concurrency, task_type="desc")

    model_to_desc: dict = {}
    for mc, data in zip(unique_models, raw_results):
        if "error" in data:
            fallback_row = rows_df.iloc[model_repr[mc]]
            model_to_desc[mc] = rule_based_short_desc(fallback_row)
        else:
            bullets = data.get("bullets", [])
            model_to_desc[mc] = "\n".join(f"• {b}" for b in bullets[:3])

    # Step 3: fan out — every SKU row gets its model's description
    descs = []
    for _, row in rows_df.iterrows():
        mc = str(row.get("model_code", "")).strip()
        if mc and mc in model_to_desc:
            descs.append(model_to_desc[mc])
        else:
            descs.append(rule_based_short_desc(row))
    return descs


# =============================================================================
# BRAND MATCHING
# =============================================================================

def match_brand(raw: str, df_brands: pd.DataFrame) -> str:
    if not raw or pd.isna(raw):
        return ""
    needle = str(raw).strip().lower()
    exact  = df_brands[df_brands["brand_name_lower"] == needle]
    if not exact.empty:
        return exact.iloc[0]["brand_entry"]
    partial = df_brands[df_brands["brand_name_lower"].str.contains(needle, regex=False)]
    if not partial.empty:
        return partial.iloc[0]["brand_entry"]
    for _, brow in df_brands.iterrows():
        if brow["brand_name_lower"] in needle:
            return brow["brand_entry"]
    return str(raw).strip()


# =============================================================================
# TEMPLATE BUILDER
# =============================================================================

def build_template(
    results_df, df_cat, df_brands,
    ai_categories,          # list[(primary, additional)] — already merged with overrides
    short_descs,            # list[str] — one per row
    variation_map,          # dict[sku → variation_string]
) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Upload Template"]

    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            header_map[val] = col_idx

    hfont      = ws.cell(row=1, column=1).font
    data_font  = Font(name=hfont.name or "Calibri", size=hfont.size or 11)
    data_align = Alignment(vertical="center")

    for i, (_, src_row) in enumerate(results_df.iterrows()):
        row_idx  = i + 2
        row_data = {}

        # Standard fields
        for master_col, tmpl_col in MASTER_TO_TEMPLATE.items():
            val = src_row.get(master_col, "")
            if pd.notna(val) and str(val).strip() not in ("", "nan"):
                row_data[tmpl_col] = str(val).strip()

        # Brand
        raw_brand = src_row.get("brand_name", "")
        if pd.notna(raw_brand) and str(raw_brand).strip():
            row_data["Brand"] = match_brand(str(raw_brand), df_brands)

        # Category (already resolved — ai_categories contains merged overrides)
        if ai_categories and i < len(ai_categories):
            primary, secondary = ai_categories[i]
        else:
            primary, secondary = keyword_match_category(src_row, df_cat)
        if primary:
            row_data["PrimaryCategory"]    = primary
        if secondary:
            row_data["AdditionalCategory"] = secondary

        # Variation
        sku = str(src_row.get("sku_num_sku_r3", "")).strip()
        row_data["variation"] = variation_map.get(sku, "size")

        # Short description
        if short_descs and i < len(short_descs) and short_descs[i]:
            row_data["short_description"] = short_descs[i]

        # Write cells
        for tmpl_col, value in row_data.items():
            if tmpl_col in header_map:
                cell           = ws.cell(row=row_idx, column=header_map[tmpl_col])
                cell.value     = value
                cell.font      = data_font
                cell.alignment = data_align

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# SIDEBAR
# =============================================================================

with st.sidebar:
    st.header("📂 Master Data")
    uploaded_master = st.file_uploader("Working file (.xlsx or .csv)", type=["xlsx", "csv"])

    st.markdown("---")
    st.header("🧠 Category Matching")
    use_ai_matching = st.toggle(
        "AI matching (Groq)",
        value=False,
        help="OFF = fast vectorised keyword/TF-IDF. ON = TF-IDF shortlist + Groq LLM rerank.",
    )

    if use_ai_matching:
        if not GROQ_AVAILABLE:
            st.error("Install groq: `pip install groq`")
            use_ai_matching = False
        else:
            st.markdown('<span class="ai-badge">AI MODE ON</span>', unsafe_allow_html=True)
            show_key     = st.checkbox("👁 Show key while typing", value=False)
            groq_api_key = st.text_input(
                "Groq API key",
                type="default" if show_key else "password",
                value=os.environ.get("GROQ_API_KEY", ""),
                placeholder="Paste your gsk_... key here",
            )
            if groq_api_key and not groq_api_key.startswith("gsk_"):
                st.warning("Groq keys usually start with `gsk_` — double-check.")
            st.caption("Free key at [console.groq.com](https://console.groq.com)")
            groq_model  = st.selectbox(
                "Model",
                ["llama-3.1-8b-instant", "llama-3.3-70b-versatile", "mixtral-8x7b-32768"],
                index=0,
                help="8b = fastest & free. 70b = most accurate.",
            )
            shortlist_k = st.slider("Shortlist size (candidates/product)", 10, 50, 30)
            concurrency = st.slider("Parallel Groq requests", 1, 30, 10)
            st.markdown("---")
            ai_short_desc = st.toggle(
                "AI short descriptions (Groq)",
                value=True,
                help="Use Groq to generate 3 polished bullet points per product. OFF = instant rule-based bullets.",
            )
    else:
        st.markdown('<span class="kw-badge">KEYWORD MODE</span>', unsafe_allow_html=True)
        st.caption("Instant vectorised TF-IDF keyword matching. No API key needed.")
        groq_api_key  = ""
        groq_model    = "llama-3.1-8b-instant"
        shortlist_k   = 30
        concurrency   = 10
        ai_short_desc = False

    st.markdown("---")
    st.header("🔎 Search Fields")
    search_fields = st.multiselect(
        "Match terms against",
        ["sku_num_sku_r3", "model_code", "model_label", "product_name", "Jumia SKU", "bar_code"],
        default=["sku_num_sku_r3"],
    )
    st.markdown("---")
    show_images = st.checkbox("Show product images", value=True)
    max_images  = st.slider("Max images per product", 1, 11, 5)


# =============================================================================
# LOAD REFERENCE DATA  (always from disk)
# =============================================================================

try:
    ref_bytes = open(DECA_CAT_PATH, "rb").read()
    st.sidebar.success("✅ deca_cat.xlsx loaded")
except FileNotFoundError:
    ref_bytes = None
    st.sidebar.error(f"⚠️ `{DECA_CAT_PATH}` not found. Place it alongside app.py and restart.")

if ref_bytes:
    df_cat, df_brands = load_reference_data(ref_bytes)
    st.sidebar.success(f"✅ {len(df_cat):,} categories · {len(df_brands)} brands")
    leaves, vectorizer, tfidf_matrix, path_to_export = build_tfidf_index(ref_bytes)
else:
    df_cat = df_brands = leaves = vectorizer = tfidf_matrix = path_to_export = None


# =============================================================================
# LOAD MASTER DATA
# =============================================================================

master_bytes = None
is_csv       = True

if uploaded_master:
    master_bytes = uploaded_master.read()
    is_csv       = uploaded_master.name.endswith(".csv")
    df_master    = load_master(master_bytes, is_csv)
    st.sidebar.success(f"✅ {len(df_master):,} product rows loaded")
else:
    loaded = False
    for path, csv in [(MASTER_PATH, True), (MASTER_PATH.replace(".csv", ".xlsx"), False)]:
        try:
            master_bytes = open(path, "rb").read()
            is_csv       = csv
            df_master    = load_master(master_bytes, csv)
            st.sidebar.info(f"📋 Bundled master · {len(df_master):,} rows")
            loaded = True
            break
        except FileNotFoundError:
            continue
    if not loaded:
        st.error("No master file found. Upload one in the sidebar.")
        st.stop()

# Pre-compute variation map for the whole master (cached)
variation_map = build_variation_map(master_bytes, is_csv) if master_bytes else {}

img_cols_present = [c for c in IMAGE_COLS if c in df_master.columns]
data_cols        = [c for c in df_master.columns if c not in img_cols_present]


# =============================================================================
# SEARCH
# =============================================================================

def search(q: str) -> pd.DataFrame:
    mask = pd.Series(False, index=df_master.index)
    for field in search_fields:
        if field not in df_master.columns:
            continue
        if field == "sku_num_sku_r3":
            # Exact match — prevents e.g. "4271703" matching "42717030"
            mask |= df_master[field].fillna("").str.strip() == q.strip()
        else:
            mask |= df_master[field].fillna("").str.lower().str.contains(q.lower(), regex=False)
    return df_master[mask].copy()


# =============================================================================
# INPUT TABS
# =============================================================================

tab1, tab2 = st.tabs(["📤 Upload a List", "⌨️ Manual Entry"])
queries = []

with tab1:
    uploaded_list = st.file_uploader(
        "Upload file with model numbers / product names",
        type=["xlsx", "csv", "txt"],
        help="One value per row. For Excel/CSV, values must be in column A.",
    )
    if uploaded_list:
        ext = uploaded_list.name.rsplit(".", 1)[-1].lower()
        if ext == "txt":
            queries = [l.strip() for l in uploaded_list.read().decode().splitlines() if l.strip()]
        elif ext == "csv":
            q_df    = pd.read_csv(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        else:
            q_df    = pd.read_excel(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        st.success(f"Loaded **{len(queries)}** search terms")

with tab2:
    manual = st.text_area(
        "Enter one SKU number per line",
        height=160,
        placeholder="4273417\n4273418\n4273423",
    )
    if manual.strip():
        queries = [q.strip() for q in manual.strip().splitlines() if q.strip()]


# =============================================================================
# RESULTS
# =============================================================================

if queries:
    st.markdown("---")
    all_result_frames = []
    no_match          = []

    for q in queries:
        res = search(q)
        if res.empty:
            no_match.append(q)
        else:
            res.insert(0, "Search Term", q)
            all_result_frames.append((q, res))

    if no_match:
        st.warning(f"No matches found for: **{', '.join(no_match)}**")

    if all_result_frames:
        total_rows = sum(len(r) for _, r in all_result_frames)
        st.success(f"**{total_rows} rows** matched across **{len(all_result_frames)}** query(ies)")

        combined = pd.concat([r for _, r in all_result_frames], ignore_index=True)

        # ── 1. Category matching ───────────────────────────────────────────────
        ai_categories = None

        if df_cat is not None and use_ai_matching and groq_api_key:
            n   = len(combined)
            est = max(2, n // concurrency + 2)
            with st.spinner(f"🤖 AI category matching {n} products (~{est}s)…"):
                try:
                    ai_categories = ai_match_categories(
                        combined, leaves, vectorizer, tfidf_matrix, path_to_export,
                        groq_api_key, groq_model, shortlist_k, concurrency,
                    )
                    st.success(f"✅ AI matched {n} products")
                except Exception as e:
                    st.error(f"Groq category error: {e}")
                    use_ai_matching = False
        elif df_cat is not None and use_ai_matching and not groq_api_key:
            st.warning("Enter your Groq API key in the sidebar to use AI matching.")
            use_ai_matching = False

        # ── 2. Short descriptions ──────────────────────────────────────────────
        short_descs = None

        if use_ai_matching and ai_short_desc and groq_api_key:
            with st.spinner(f"✍️ Generating AI short descriptions ({len(combined)} products)…"):
                try:
                    short_descs = ai_short_descriptions(combined, groq_api_key, groq_model, concurrency)
                    st.success("✅ Short descriptions generated")
                except Exception as e:
                    st.error(f"Short desc error: {e}")
                    short_descs = None

        if short_descs is None:
            # Rule-based fallback (instant)
            short_descs = [rule_based_short_desc(row) for _, row in combined.iterrows()]

        # ── 3. Category & Brand preview with manual override ───────────────────
        if df_cat is not None:
            mode_label = "🤖 AI" if (use_ai_matching and ai_categories) else "🔑 Keyword"
            with st.expander(f"{mode_label} — Category, Variation & Description Preview", expanded=False):

                all_export_cats             = sorted(df_cat["export_category"].dropna().unique().tolist())
                all_export_cats_with_blank  = ["(auto)"] + all_export_cats

                if "cat_overrides" not in st.session_state:
                    st.session_state.cat_overrides = {}

                st.markdown(
                    "**Override categories per row** — choose `(auto)` to keep the matched value."
                )
                st.markdown("---")

                hc1, hc2, hc3, hc4, hc5, hc6 = st.columns([2, 3, 3, 1, 1, 2])
                hc1.markdown("**Product**")
                hc2.markdown("**Primary Category**")
                hc3.markdown("**Additional Category**")
                hc4.markdown("**Variation**")
                hc5.markdown("**Method**")
                hc6.markdown("**Short Description**")

                for i, (_, prow) in enumerate(combined.iterrows()):
                    if use_ai_matching and ai_categories:
                        auto_prim, auto_addl = ai_categories[i]
                    else:
                        auto_prim, auto_addl = keyword_match_category(prow, df_cat)

                    override = st.session_state.cat_overrides.get(i, {})
                    sku      = str(prow.get("sku_num_sku_r3", "")).strip()
                    var_val  = variation_map.get(sku, "size")
                    sd_val   = short_descs[i] if short_descs else ""

                    c1, c2, c3, c4, c5, c6 = st.columns([2, 3, 3, 1, 1, 2])
                    c1.markdown(
                        f"**{sku}**  \n"
                        f"{str(prow.get('product_name', ''))[:50]}"
                    )

                    cur_prim = override.get("primary", auto_prim)
                    try:
                        prim_idx = all_export_cats_with_blank.index(cur_prim)
                    except ValueError:
                        prim_idx = 0
                    new_prim = c2.selectbox(
                        f"Primary #{i}", all_export_cats_with_blank,
                        index=prim_idx, label_visibility="collapsed", key=f"prim_{i}",
                    )

                    cur_addl = override.get("additional", auto_addl)
                    try:
                        addl_idx = all_export_cats_with_blank.index(cur_addl)
                    except ValueError:
                        addl_idx = 0
                    new_addl = c3.selectbox(
                        f"Additional #{i}", all_export_cats_with_blank,
                        index=addl_idx, label_visibility="collapsed", key=f"addl_{i}",
                    )

                    c4.markdown(f"`{var_val}`")

                    if new_prim != "(auto)" or new_addl != "(auto)":
                        st.session_state.cat_overrides[i] = {
                            "primary":    auto_prim if new_prim == "(auto)" else new_prim,
                            "additional": auto_addl if new_addl == "(auto)" else new_addl,
                        }
                    elif i in st.session_state.cat_overrides:
                        del st.session_state.cat_overrides[i]

                    badge = "🖊️ Manual" if i in st.session_state.cat_overrides else (
                        "🤖 AI" if (use_ai_matching and ai_categories) else "🔑 Keyword"
                    )
                    c5.markdown(f"`{badge}`")
                    c6.markdown(sd_val.replace("\n", "  \n") if sd_val else "_—_")

                st.markdown("---")
                st.caption(
                    f"Total rows: {len(combined)} · "
                    f"Overrides: {len(st.session_state.cat_overrides)}"
                )

        # ── Download buttons ───────────────────────────────────────────────────
        col_dl1, col_dl2 = st.columns(2)

        with col_dl1:
            raw_out = io.BytesIO()
            with pd.ExcelWriter(raw_out, engine="openpyxl") as writer:
                combined.to_excel(writer, index=False, sheet_name="Results")
            st.download_button(
                "⬇️ Download Raw Results (.xlsx)",
                data=raw_out.getvalue(),
                file_name="decathlon_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col_dl2:
            if df_cat is None:
                st.warning("deca_cat.xlsx not loaded — template download unavailable.")
            else:
                try:
                    # Merge session-state overrides on top of AI/keyword categories
                    merged_cats = []
                    for i, (_, prow) in enumerate(combined.iterrows()):
                        override = st.session_state.get("cat_overrides", {}).get(i)
                        if override:
                            merged_cats.append((override["primary"], override["additional"]))
                        elif use_ai_matching and ai_categories:
                            merged_cats.append(ai_categories[i])
                        else:
                            merged_cats.append(keyword_match_category(prow, df_cat))

                    tpl_bytes = build_template(
                        combined, df_cat, df_brands,
                        ai_categories=merged_cats,
                        short_descs=short_descs,
                        variation_map=variation_map,
                    )
                    mode_icon = "🤖" if (use_ai_matching and ai_categories) else "🔑"
                    st.download_button(
                        f"{mode_icon} Download Filled Upload Template (.xlsx)",
                        data=tpl_bytes,
                        file_name="decathlon_upload_template_filled.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                except FileNotFoundError:
                    st.warning(
                        "Template file not found. "
                        "Place `product-creation-template.xlsx` in the app folder."
                    )

        st.markdown("---")

        # ── Per-query result cards ─────────────────────────────────────────────
        for q, res in all_result_frames:
            with st.expander(f"🔍 **{q}**  —  {len(res)} row(s)", expanded=True):
                show_cols = ["Search Term"] + [c for c in data_cols if c in res.columns]
                st.dataframe(
                    res[show_cols],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "keywords":     st.column_config.TextColumn("keywords",     width="large"),
                        "product_name": st.column_config.TextColumn("product_name", width="large"),
                        "designed_for": st.column_config.TextColumn("designed_for", width="large"),
                    },
                )

                cats = set()
                for _, row in res.iterrows():
                    for col in ["department_label", "nature_label", "family", "type"]:
                        val = row.get(col, "")
                        if pd.notna(val) and str(val).strip():
                            cats.add(str(val).strip())
                if cats:
                    tags = " ".join(f'<span class="tag">{c}</span>' for c in sorted(cats))
                    st.markdown(f"**Categories & Types:** {tags}", unsafe_allow_html=True)

                if show_images and img_cols_present:
                    first_row = res.iloc[0]
                    img_urls  = [
                        str(first_row[c]) for c in img_cols_present
                        if pd.notna(first_row.get(c))
                        and str(first_row.get(c, "")).startswith("http")
                    ][:max_images]
                    if img_urls:
                        st.markdown("**🖼 Product Images**")
                        cols = st.columns(len(img_urls))
                        for i, url in enumerate(img_urls):
                            try:
                                resp = requests.get(url, timeout=6)
                                img  = Image.open(io.BytesIO(resp.content))
                                cols[i].image(
                                    img,
                                    caption="Main" if i == 0 else f"View {i}",
                                    use_container_width=True,
                                )
                            except Exception:
                                cols[i].markdown(f"[🔗 Image {i+1}]({url})")
else:
    st.info("👆 Upload a list or type search terms above to get started.")

st.markdown("---")
st.caption("Decathlon Product Lookup · Powered by your Decathlon working file")
